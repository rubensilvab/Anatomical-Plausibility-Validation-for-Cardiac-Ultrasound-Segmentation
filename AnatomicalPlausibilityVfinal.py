# -*- coding: utf-8 -*-
"""
Created on Wed Jul 31 16:58:38 2024

@author: RubenSilva
"""


import nibabel as nib
import os
import matplotlib.pyplot as plt 
import numpy as np 
import cv2
import cc3d
from skimage import measure

Name_classes={0: 'Background', 1: 'Endocardium', 2:'Myocardium', 3: 'left atrium'}

def SeeSpecificStructure(structure, data):
    # Create a boolean mask where the elements are equal to the structure
    mask = (data == structure)
    
    if structure ==0:
        mask=mask.astype(int)
        return np.logical_not(mask).astype(int)
    else:
        # Use the mask to zero out elements that are not part of the structure
        result = np.where(mask, data, 0)
        return result

def BinaryzeAllMask(data):
    result=np.copy(data)
    mask = (data >=1)
    result[np.where(mask)] = 1
    
    return result

"""Criteria 1 and 2"""

def detect_holes(image):
    # Convert the image to grayscale if it is not already
    image = np.transpose(image, (1, 0))
    image = BinaryzeAllMask(image).astype(np.uint8)
    #print(np.unique(image))
    mask_fill=image.copy()
    filled_img = np.zeros_like(mask_fill)
    holes_img = np.zeros_like(mask_fill)
    
    
    # Find contours
    contours, hierarchy= cv2.findContours(mask_fill, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    # Draw contours on black image
    contour_img = np.zeros_like(mask_fill)
    
    cv2.drawContours(contour_img, contours, -1, 1, 2)
    
    for contour in contours:
                cv2.fillPoly(filled_img, contours, 1)
    
    #print(np.unique(filled_img))
    holes_img=filled_img-image
    
    contours = measure.find_contours(holes_img, 0.5)
    n_holes=len(contours)
    
    if np.sum(holes_img)>1:
        #print('There are holes in the segmentation, no plausible anatomy')
        holes=True

    else: 
        holes=False
        n_holes=0
        
    return filled_img,contour_img,holes_img,holes,n_holes

"""Criteria 3"""

def connected_components(image,n):
     # Get a labeling of the k largest objects in the image.
     # The output will be relabeled from 1 to N.
     labels_out, N = cc3d.largest_k(
       image, k=n, 
       connectivity=4, delta=0,
       return_N=True,
     )
    
     labels_out=labels_out.astype(np.uint8)
     # labels_out[labels_out==2]=0
     # plt.imshow(labels_out)
     # plt.show()
     # print(np.unique(labels_out), np.sum(labels_out))
     for i in range(len(np.unique(labels_out))-1):
         n_points=len(np.where(labels_out==i+1)[0])
         #print(n_points, 'classe',i+1,np.unique(labels_out))
         if n_points<15:
             N=N-1

     return labels_out, N
 
""" Criteria 4- calculate ratio between structures (perimeters)"""


def RatioPerInt_new(contours_1,contours_2,inter12):
    
    """Perimeter in number of points"""
    
    per_1=len(contours_1)
    per_2=len(contours_2)
    Pxinter=len(inter12)
    
    ratio_1=Pxinter/per_1
    ratio_2=Pxinter/per_2
    
    return ratio_1,ratio_2

def DefineContourSkimage(mask):
    
   # Compute the contours of the mask to be able to analyze each part individually
   contours = measure.find_contours(mask, 0.5)
   
   # Initialize arrays to store the curvature information for each edge pixel
   edge_pixels = []
   
   # Iterate over each contour
   for contour in contours:
       # Iterate over each point in the contour
       if (contour.shape[0]>30):
           for i, point in enumerate(contour):    
            # Store curvature information and corresponding edge pixel
            edge_pixels.append(point)
     
   # Convert lists to numpy arrays for further processing
   edge_pixels = np.array(edge_pixels)            
   
   return edge_pixels  

def dist2points(p1,p2):
    diff=p2-p1
    #print(diff)
    power=np.sum(diff**2)
    #print(power)
    dist=np.sqrt(power)
    
    return dist

from scipy.spatial import KDTree

def FindCommonPoints_new(c1, c2, delta):
    # Build KD-tree for the second set of points
    tree = KDTree(c2)
    
    CommonPointsc1 = []
    CommonPointsc2 = []
    
    for point in c1:
        # Query the KD-tree for neighbors within delta
        indices = tree.query_ball_point(point, delta)
        for idx in indices:
            CommonPointsc1.append(point)
            CommonPointsc2.append(c2[idx])
    
    return np.array(CommonPointsc1), np.array(CommonPointsc2)

def PointsCloseBorder(c1,image,delta):
    y,x=image.shape
    
    pointsborder=[]
    ind=[]
    for i,point in enumerate(c1):
        p_border_r,p_border_l=np.copy(point),np.copy(point)
        p_border_r[1],p_border_l[1]=x,0
        
        if dist2points(point,p_border_r)<delta or dist2points(point,p_border_l)<delta :
           pointsborder.append(point)
           ind.append(i)
           
    return pointsborder, ind         
           
          
def RemovePoints(c,inter):
    #c=list(c)
    #inter=list(inter)
    # Create a boolean mask for points to be removed
    mask = np.ones(len(c), dtype=bool)
    
    for point in inter:
        mask &= ~np.all(c == point, axis=1)
    
    # Apply the mask to filter out points
    filtered_points = c[mask]
    #plt.scatter(filtered_points[:, 1], filtered_points[:,0])
    #plt.show()
    
    return filtered_points         
      

def FindFirstPoint(contour,first_ind):
    #first_ind: first indice of the point which we want to find if this is the first in the open contour 
    ind=first_ind
    distance=0
    while distance<3:
        distance=dist2points(contour[ind],contour[ind-1])
        ind=ind-1 
    return ind+1    

def order_contour_points(points,n_class=1):
    
    if n_class==2:
        #top righ index
        first_index = FindFirstPoint(points,0)
        first_point=points[first_index]
    
    elif n_class=='centerline':
        
        first_index=np.argmax(points[:,1])
        first_point = points[first_index]
        
    else:
        # Find the bottom-left point
        first_index = np.argmin(points[:, 0] + points[:, 1])
        first_index = FindFirstPoint(points,first_index)
        first_point = points[first_index]
    
    # Initialize ordered points with the top-left point
    ordered_points = [first_point]
    
    # Remove the top-left point from the list of points
    points = np.delete(points, first_index, axis=0)
    
    while points.shape[0] > 0:
        # Find the closest point to the last point in ordered_points
        last_point = ordered_points[-1]
        distances = np.linalg.norm(points - last_point, axis=1)
        nearest_index = np.argmin(distances)
        nearest_point = points[nearest_index]
        
        # Add the nearest point to ordered_points and remove it from points
        ordered_points.append(nearest_point)
        points = np.delete(points, nearest_index, axis=0)
    
    return np.array(ordered_points)

def AddDistancePoint(contour):

 distance=np.zeros(contour.shape[0])    
 for i in range(contour.shape[0]):
     if i==0:
         distance[i]=0
     else:
         distance[i]=dist2points(contour[i],contour[i-1])+distance[i-1] 
         
 return distance    


from scipy import interpolate
"""Make interpolation in order to put the points equidistance"""

def ContourInterpolation(contour,retrieve_all_plots,name):
      
  distance=AddDistancePoint(contour)
  
  fx = interpolate.interp1d(distance, contour[:,1])
  fy= interpolate.interp1d(distance, contour[:,0])
  
  dist_max=distance[-1]
  dist_p_points=1
  
  distances= np.arange(0,dist_max,dist_p_points)
  
  xnew = fx(distances)
  ynew=  fy(distances)
  
  contourfit=np.transpose(np.array([ynew,xnew]))
  
  if retrieve_all_plots:
      plt.title(name+'Contour after interpolation')
      plt.scatter(contourfit[:, 1], contourfit[:, 0])
      plt.show()
  
  
  return contourfit


def derivative(x):
    derivative=[]
    for i in range(len(x)-1):
        derivative.append(x[i+1]-x[i])
    return derivative

#coeffs = np.polyfit(np.arange(0, len(points[:,x]-1), 1), points[:,x], 10)


"""Criteria 5: Thickness"""

from sklearn.decomposition import PCA

def CropX(c, th):
    points_remove=[]
    for point in c:   
        if point[1]< th:
            points_remove.append(point)
            
    filtered_c=RemovePoints(c, points_remove)
    
    return filtered_c  

    
def RemoveSegmentsMyo(c_pca, c_myo_all,c_myo,c_endo ,retrieve_all_plots,tolerance=25, endo=0):
      
    pca = PCA(n_components=2)
    pca.fit(c_pca-np.mean(c_pca,axis=0))
 
    pca_comp=pca.components_ 
          
    if pca_comp[1,1]>0:   
        pca_comp=-pca_comp
    
    """Myocardium rotated"""
    c_myo_all=c_myo_all-np.mean(c_pca,axis=0) 
    c_myo_all =np.matmul(c_myo_all,pca_comp.T)
    
    c_myo_r= c_myo-np.mean(c_pca,axis=0)  
    c_myo_r =np.matmul(c_myo_r,pca_comp.T)
    
    """Endocardium rotated"""
    
    c_endo_r= c_endo-np.mean(c_pca,axis=0)  
    c_endo_r =np.matmul(c_endo_r,pca_comp.T)
    
    """Apply threshold"""
      
    ratio=0.070294451
    h_img= np.max(c_myo_all[:,1])-np.min(c_myo_all[:,1])
    h_crop= ratio*h_img + tolerance
    
    c_myo_r=CropX(c_myo_r, np.min(c_myo_all[:,1])+h_crop)
    c_endo_r=CropX(c_endo_r, np.min(c_myo_all[:,1])+h_crop)
    
    LV_width= np.max(c_endo_r[:,0])-np.min(c_endo_r[:,0])
    
    """Myocardium cropped in original position"""  
  
    c_myo_r=np.matmul(c_myo_r,pca_comp)
    c_myo_r=c_myo_r+ np.mean(c_pca,axis=0)
    
    c_endo_r=np.matmul(c_endo_r,pca_comp)
    c_endo_r=c_endo_r+ np.mean(c_pca,axis=0)
    
    if retrieve_all_plots:
        plt.title('Contour Myo and Endo: after Myo segments removed')
        plt.scatter(c_myo_r[:,1],c_myo_r[:,0])
        plt.scatter(c_endo_r[:,1],c_endo_r[:,0])
        # Set equal scaling
        plt.axis('equal')
        plt.show()
    
    return c_myo_r, c_endo_r, LV_width


def measure_thickness(c_myo,c_endo,retrieve_all_plots): 

    # Build KDTree for contour2
    tree = KDTree(c_endo)

    # Find the minimal distance from each point in contour1 to contour2
    distances, _ = tree.query(c_myo)
    
    if retrieve_all_plots:
        plt.title('Myo and Endo used to calculate thickness (check if it is correct)')
        plt.scatter(c_myo[:,1],c_myo[:,0])
        plt.scatter(c_endo[:,1],c_endo[:,0])
        plt.show()
    
    min_thickness = np.min(distances)
    max_thickness = np.max(distances)
    avg_thickness = np.mean(distances)
    
    
    return max_thickness, min_thickness, avg_thickness, distances 

"""Criteria 7: Curvature"""

def CurvatureValues_new(contour,window_size,retrieve_all_plots,name):
   
   # Initialize arrays to store the curvature information for each edge pixel
   curvature_values = []
   
    # Iterate over each point in the contour
   for i, point in enumerate(contour):
    # Compute the curvature for the point
    # We set the window size to 1/5 of the whole contour edge. Adjust this value according to your specific task
    #window_size=int(len(contour)/window_size_ratio)
    curvature = compute_curvature(point, i, contour, window_size)
     
     # Store curvature information and corresponding edge pixel
    curvature_values.append(curvature)
    
   # Convert lists to numpy arrays for further processing
   curvature_values = np.array(curvature_values)
   
   if retrieve_all_plots:
       
       plt.scatter(contour[:, 1], contour[:, 0],c=curvature_values, cmap='jet',vmin=-0.12, vmax=0.12)
       plt.colorbar(label='Curvature')
       plt.title(name+" Curvature of Edge Pixels")
       plt.show()

   return  curvature_values  

def compute_curvature(point, i, contour, window_size):
    # Compute the curvature using polynomial fitting in a local coordinate system
    if (((i - window_size // 2)<0 )&(dist2points(contour[0], contour[-1])<1)):
        
        start=i - window_size // 2
        end=-1
        contour_before=contour[start:end]
        contour_after=contour[0:i + window_size // 2 + 1]
        neighborhood= np.concatenate((contour_before, contour_after))
     
    elif ((i + window_size // 2> len(contour)) &(dist2points(contour[0], contour[-1])<1)):
        
        start=i - window_size // 2
        end=-1
        contour_before=contour[start:end]
        contour_after=contour[0:i + window_size // 2 + 1 - len(contour)]
        neighborhood= np.concatenate((contour_before, contour_after))
    
        
    else:
            
        # Extract neighboring edge points
        start = max(0, i - window_size // 2)
        end = min(len(contour), i + window_size // 2 + 1)
        neighborhood = contour[start:end]
    
    # #Extract neighboring edge points
    # start = max(0, i - window_size // 2)
    # end = min(len(contour), i + window_size // 2 + 1)
    # neighborhood = contour[start:end]

    # Extract x and y coordinates from the neighborhood
    x_neighborhood = neighborhood[:, 1]
    y_neighborhood = neighborhood[:, 0]

    # Compute the tangent direction over the entire neighborhood and rotate the points
    tangent_direction_original = np.arctan2(np.gradient(y_neighborhood), np.gradient(x_neighborhood))
    tangent_direction_original.fill(tangent_direction_original[len(tangent_direction_original)//2])

    # Translate the neighborhood points to the central point
    translated_x = x_neighborhood - point[1]
    translated_y = y_neighborhood - point[0]


    # Apply rotation to the translated neighborhood points
    # We have to rotate the points to be able to compute the curvature independent of the local orientation of the curve
    rotated_x = translated_x * np.cos(-tangent_direction_original) - translated_y * np.sin(-tangent_direction_original)
    rotated_y = translated_x * np.sin(-tangent_direction_original) + translated_y * np.cos(-tangent_direction_original)

    # Fit a polynomial of degree 2 to the rotated coordinates
    coeffs = np.polyfit(rotated_x, rotated_y, 2)


    # You can compute the curvature using the formula: curvature = |d2y/dx2| / (1 + (dy/dx)^2)^(3/2)
    # dy_dx = np.polyval(np.polyder(coeffs), rotated_x)
    # d2y_dx2 = np.polyval(np.polyder(coeffs, 2), rotated_x)
    # curvature = np.abs(d2y_dx2) / np.power(1 + np.power(dy_dx, 2), 1.5)

    # We compute the 2nd derivative in order to determine whether the curve at the certain point is convex or concave
    curvature = np.polyval(np.polyder(coeffs, 2), rotated_x)

    # Return the mean curvature for the central point
    return np.mean(curvature)       

import openpyxl


def ExcelRowResult(path,criteria_array,patient):
    
    isExist = os.path.exists(path)
    if not isExist:                         
        # Create a new directory because it does not exist 
        os.makedirs(path)  
    os.chdir(path)   

    name='AnatomicalPlausibiityResults'        
    filename = str(name)+'.xlsx'
    
    print('Saving evaluation metrics to excel...')
    # Check if the file already exists
    if not os.path.isfile(filename):
        # Create a new workbook object
        book = openpyxl.Workbook()
        # Select the worksheet to add data to
        sheet = book.active
        # Add a header row to the worksheet
        sheet.append(['Patient', 'Criteria 1', 'Criteria 2', 'Criteria 3', 'Criteria 4', 'Criteria 5', 'Criteria 6','Criteria 7'])
    else:
        # Open an existing workbook
        book = openpyxl.load_workbook(filename)
    
        # Select the worksheet to add data to
        sheet = book.active
    
    # Append a new row of data to the worksheet
    sheet.append([patient, str(criteria_array[0]),str(criteria_array[1]), str(criteria_array[2]),str(criteria_array[3]), str(criteria_array[4]),str(criteria_array[5]),str(criteria_array[6])])
    #print([patient, slic, label, pred])
    # Save the workbook to a file
    
    book.save(filename)


def ExcelRowThresolds(path,cv_metric1,cv_metric2,cv_metric3,RatioEndoBckg,RatioMyoLA,RatioThk_minmax,RatioThk_avgWLV,patient):
    
    isExist = os.path.exists(path)
    if not isExist:                         
        # Create a new directory because it does not exist 
        os.makedirs(path)  
    os.chdir(path)   

    name='MetricsResult'        
    filename = str(name)+'.xlsx'
    
    print('Saving evaluation metrics to excel...')
    # Check if the file already exists
    if not os.path.isfile(filename):
        # Create a new workbook object
        book = openpyxl.Workbook()
        # Select the worksheet to add data to
        sheet = book.active
        # Add a header row to the worksheet
        sheet.append(['Patient', 'Curvature Endo Max', 'Curvature Endo Min', 'Curvature Myo Max', 'Curvature Myo Min', 'Curvature LA Max', 'Curvature LA Min','Ratio Int-EndoBckg/Bckg','Ratio Int-EndoBckg/Endo','Ratio Int-MyoLA/Myo','Ratio Int-MyoLA/LA','Ratio Min/Max Myo Thikness','Ratio Avg/WidthLV Myo Thikness'])
    else:
        # Open an existing workbook
        book = openpyxl.load_workbook(filename)
    
        # Select the worksheet to add data to
        sheet = book.active
    
    # Append a new row of data to the worksheet
    sheet.append([patient, cv_metric1[1],cv_metric1[0], cv_metric2[1],cv_metric2[0], cv_metric3[1],cv_metric3[0],RatioEndoBckg[0],RatioEndoBckg[1],RatioMyoLA[0],RatioMyoLA[1],RatioThk_minmax,RatioThk_avgWLV])
    #print([patient, slic, label, pred])
    # Save the workbook to a file
    
    book.save(filename)

def FindNumberContours(c,delta):
    contours=[]
    indices=[0]
    
    for i in range(len(c)-1):
        if dist2points(c[i],c[i+1])>delta:
            indices.append(i)
            
    indices.append(-1)           
            
    if len(indices)>2:
        for i in range(len(indices)-1):
            if i==0:
                contour=c[indices[i]:indices[i+1]+1]
                contours.append(contour)
            elif (i==(len(indices)-2)):
                contour=c[indices[i]+1:]
                contours.append(contour)
            else:
                contour=c[indices[i]+1:indices[i+1]+1]
                contours.append(contour)
    
    else:
        
        contours.append(c)
        
    """If the contour is too small, remove it"""
    
    for i in reversed(range(len(contours))):
        if len(contours[i]) < 15:
            del contours[i]  
    
    return contours    
        
def CalculateCurvatureMulti_Cont(c,window_size,retrieve_all_plots,name):
    
   curvature_values=[]
   for i in range(len(c)):
     
    curvature_valuei=CurvatureValues_new(c[i],window_size,retrieve_all_plots,name+' '+str(i+1))
    curvature_values.append(curvature_valuei)
   
   # Flatten the list of contours into a single list of curvature values
   flatten_curvature = [curvature for contour in curvature_values for curvature in contour]
   flattened_contours = [point for contour in c for point in contour]   
   
   contours=np.array(flattened_contours)
   
   if len(c)!=0:
       if retrieve_all_plots:
           plt.title(name+" Curvature of Edge Pixels")
           plt.scatter(contours[:,1],contours[:,0],c=flatten_curvature, cmap='jet',vmin=-0.12, vmax=0.12)
           plt.colorbar(label='Curvature')
           plt.show()   
   else:
       print('Does not have contour')
   
   return flatten_curvature 
          

def CheckPlausibility(image,n_classes,cardiac_ph,path,patient,retrieve_all_plots):
    
    if retrieve_all_plots:
        plt.imshow(image, cmap='gray')
        plt.show()
        
    n_holes_class=[]
    holes_imgs=[]
    clean_img=[]
    cr1,cr2,cr3,cr4,cr5,cr6, cr7= np.zeros((3,)),np.zeros((2,)),np.zeros((3,)),np.zeros((2,)),np.zeros((1,)),np.zeros((1,)),np.zeros((3,))
    
        
    for i in range(n_classes):
         
        data_n=SeeSpecificStructure(i+1, image)
        data_b_bi=BinaryzeAllMask(data_n)
       
        if retrieve_all_plots:
            plt.imshow(data_b_bi, cmap='gray')
            plt.show()
        
        filled_img,_,holes,_,n_holes=detect_holes(data_b_bi)
        n_holes_class.append(n_holes)
        holes_imgs.append(holes)
        
        #print(n_holes)
        if n_holes>0:
            print('There are',n_holes,' holes in the',Name_classes[i+1],' segmentation, no plausible anatomy')
            cr1[i]=1
             
        _,Ni=connected_components(data_b_bi,3)
        pp_img,_=connected_components(filled_img,1)
        clean_img.append(pp_img)
        
        if Ni>1:
            print('There are', Ni,' disconnected componentes in the class', Name_classes[i+1],', no anatomically plausible')
            cr3[i]=1 
        
    """Common criteria (not need to iterate for each class)"""
    
    # Presence of holes between LV and MYO
    mask_myo_lv=np.copy(image)
    mask_myo_lv[mask_myo_lv==3]=0
    data_myo_lv=BinaryzeAllMask(mask_myo_lv)
    
    _,_,total_holes,_,holes_myo_lv=detect_holes(data_myo_lv)
    holes_int=holes_myo_lv-n_holes_class[0]-n_holes_class[1]
    
    if (holes_int)>0:
        print('There are',holes_int,' holes between LV and MYO, no plausible anatomy')
        cr2[0]=1
        #Add intersection holes to Myocardium to the contours be considered as intersection LV-MYO and be removed
        holes_myo_lv= total_holes-holes_imgs[0]-holes_imgs[1]
        clean_img[1]=clean_img[1]+holes_myo_lv
        
    
    # Presence of holes between LV and LA
    mask_la_lv=np.copy(image)
    mask_la_lv[mask_la_lv==2]=0
    data_la_lv=BinaryzeAllMask(mask_la_lv)
    
    _,_,total_holes_lalv,_,n_holes_la_lv=detect_holes(data_la_lv)
    holes_int_lav=n_holes_la_lv-n_holes_class[0]-n_holes_class[2]
    
    if holes_int_lav>0:
        print('There are',holes_int_lav,' holes between LV and LA, no plausible anatomy')
        cr2[1]=1
        #Add intersection holes to la to the contours be considered as intersection LV-LA and be removed
        holes_la_lv= total_holes_lalv-holes_imgs[0]-holes_imgs[2]
        clean_img[2]=clean_img[2]+holes_la_lv
    
    """ Critera 4:Size of the perimeter by which the LV touches the background or the MYO touches the LA exceeds a certain threshold."""
    
    """Image has to be clean, without holes or disconnected components"""
    backgrd=SeeSpecificStructure(0, image) # Background
    #clean background
    filled_img,_,_,_,_=detect_holes(backgrd)
    Seebackgrd,_=connected_components(filled_img,1)
    
    SpecificSeg0=np.transpose(Seebackgrd, (1, 0)) #Background
    SpecificSeg1=np.transpose(clean_img[0], (1, 0)) # Endocardium
    SpecificSeg2=np.transpose(clean_img[1], (1, 0)) # Myocardium
    SpecificSeg3=np.transpose(clean_img[2], (1, 0)) # LA
    
    """ Find Contours"""
    contour_0=DefineContourSkimage(SpecificSeg0)
    contour_1=DefineContourSkimage(SpecificSeg1)
    contour_2=DefineContourSkimage(SpecificSeg2)
    contour_3=DefineContourSkimage(SpecificSeg3)
    
    "Interpolation in order to make all the points equidistances"
    
    contour_0=ContourInterpolation(contour_0,retrieve_all_plots,'All ')
    contour_1=ContourInterpolation(contour_1,retrieve_all_plots,'Endo ')
    contour_2=ContourInterpolation(contour_2,retrieve_all_plots,'Myo ')
    contour_3=ContourInterpolation(contour_3,retrieve_all_plots,'LA ')
    
    
    "Find intersections between contours"
    
    inter_13,inter_31=FindCommonPoints_new(contour_1,contour_3,3) # Endo and LA
    inter_21,_=FindCommonPoints_new(contour_2,contour_1,4) # Myo and Endo
    
    """Remove intersections from the contours"""
    new_contour_1=RemovePoints(contour_1,inter_13)
    new_contour_2=RemovePoints(contour_2,inter_21)
    new_contour_3=RemovePoints(contour_3,inter_31)
    
    """Order contour"""
    
    new_contour_1=order_contour_points(new_contour_1,2)
    new_contour_2=order_contour_points(new_contour_2,2) # fix problem with order Myo
    new_contour_3=order_contour_points(new_contour_3) # fix problem with order Atrium 

    """Criteria 4 (1) -Intersecion endo bckg"""

    inter_01,_=FindCommonPoints_new(new_contour_1,contour_0,1)
    ratio_IntBckg,ratio_IntLV=RatioPerInt_new(contour_0,new_contour_1,inter_01)
    #print('Ratio (Intersetion between LV and background/ Background: ',ratio_IntBckg,' and ratio.../ Endocardium: ', ratio_IntLV)    
   
    if (ratio_IntBckg> 0 or ratio_IntLV>0): 
        cr4[0]=1
        
    """Criteria 4 (2) -Intersecion MYO LA"""

    inter_23,_=FindCommonPoints_new(new_contour_2,new_contour_3,1)
    ratio_IntMyo,ratio_IntLA=RatioPerInt_new(new_contour_2,new_contour_3,inter_23)      
    #print('Ratio (Intersetion between MYO and LA/ MYO: ',ratio_IntMyo,' and ratio.../ LA: ', ratio_IntLA)    
           
    if cardiac_ph=='ES':
        th_IntMyo,th_IntLA=0.132,0.237
    else:
        th_IntMyo,th_IntLA=0.222,0.35
    
    if (ratio_IntMyo> th_IntMyo or ratio_IntLA>th_IntLA):    
        cr4[1]=1
        
    """Criterio 5 and 6- Myocardium thickness (2)"""
    
    #Crop the two parts of the myocardium
    
    """Apply thresholds to remove Myocadium bases with PCA"""
    contour_myo,contour_endo,LV_width=RemoveSegmentsMyo(inter_13, contour_2,new_contour_2,new_contour_1,retrieve_all_plots)   
    
    """Correction if myo is cropped"""   
    #Myocardium
    
    c2_points_remove,ind= PointsCloseBorder(contour_myo,image,2)
    # Remove the specified indices
    contour_myo = np.delete(contour_myo, ind, axis=0)
    
    if (retrieve_all_plots and len(ind)>0):
        plt.title('Myo with borders removed (if it is cropped)')
        plt.scatter(contour_myo[:, 1], contour_myo[:, 0])
        plt.show()
    
    """Find Myocardium thickness"""
    
    thk_max,thk_min,thk_avg,_=measure_thickness(contour_myo,contour_endo,retrieve_all_plots)
    
    ratio_minmax_thk=thk_min/thk_max
    ratio_thkmin_lv_w=thk_avg/LV_width
    
    if cardiac_ph=='ES':
        ratioMiMa_min,ratioMiMa_max,ratioMiLV_min,ratioMiLV_max=0.297,0.907,0.167,2.24
    else:
        ratioMiMa_min,ratioMiMa_max,ratioMiLV_min,ratioMiLV_max=0.268,0.846,0.144,0.623
        
        
    if (ratio_minmax_thk<ratioMiMa_min) or (ratio_minmax_thk>ratioMiMa_max):
        cr5[0]=1
    
    if (ratio_thkmin_lv_w<ratioMiLV_min) or (ratio_thkmin_lv_w>ratioMiLV_max):
        cr6[0]=1
    
    new_contour_2=contour_myo
    new_contour_2=order_contour_points(new_contour_2,2)
    
    """Correction if the LA is cropped"""
    
    c3_points_remove,ind= PointsCloseBorder(new_contour_3,image,2)
    # Remove the specified indices
    new_contour_3 = np.delete(new_contour_3, ind, axis=0)
     
    if retrieve_all_plots and len(ind)>0:
        plt.title('LA with borders removed (if it is cropped)')
        plt.scatter(new_contour_3[:, 1], new_contour_3[:, 0])
        plt.show()

    """Find if there are two or more contours"""
    
    #Atrium
    new_contours_3v1=FindNumberContours(new_contour_3,3)
    #Myocardium
    new_contours_2v1=FindNumberContours(new_contour_2,3)
    
    
    """Criteria 7: Curvature"""
    
    # Set the window size for local polynomial approximation
    window_size = 30
    
    cv_1=CurvatureValues_new(new_contour_1,window_size,retrieve_all_plots,'Endo')
    cv_2=CalculateCurvatureMulti_Cont(new_contours_2v1,window_size,retrieve_all_plots,'Myo')
    cv_3=CalculateCurvatureMulti_Cont(new_contours_3v1,window_size,retrieve_all_plots,'LA')
    
    """Calculate min, max curvatures:"""
    
    cv_metric1 = [np.nan, np.nan] if len(cv_1) == 0 else [np.min(cv_1), np.max(cv_1)]
    cv_metric2 = [np.nan, np.nan] if len(cv_2) == 0 else [np.min(cv_2), np.max(cv_2)]
    cv_metric3 = [np.nan, np.nan] if len(cv_3) == 0 else [np.min(cv_3), np.max(cv_3)]
    

    ExcelRowThresolds(path,cv_metric1,cv_metric2,cv_metric3,[ratio_IntBckg,ratio_IntLV],[ratio_IntMyo,ratio_IntLA],ratio_minmax_thk,ratio_thkmin_lv_w,patient)
    
    if cardiac_ph=='ES':
        th1max,th1min,th2max,th2min,th3max,th3min=0.461,-0.044,0.0542,-0.0279,0.1346,-0.038
    else:
        th1max,th1min,th2max,th2min,th3max,th3min=0.153,-0.049,0.0563,-0.026,0.1332,-0.063

    
    if (cv_metric1[1] > th1max) or (cv_metric1[0]< th1min): 
    
      cr7[0]=1
    
    if (cv_metric2[1] > th2max) or (cv_metric2[0]< th2min): 
      
      cr7[1]=1
    
    if (cv_metric3[1] > th3max) or (cv_metric3[0]< th3min): 
      
      cr7[2]=1  
        
    plausibility=[cr1,cr2,cr3,cr4,cr5,cr6,cr7] 
    return plausibility   



"""Critérios:
    1-(3 criteria) Presence of holes in the Left Ventricle (LV), Myocardium (MYO), and Left Atrium (LA).
    2-(2 criteria) Presence of holes between the LV and MYO, or between the LV and LA.
    3-(3 criteria) Presence of more than one LV, MYO, or LA. 
    4-(2 criteria) Size of the perimeter by which the LV touches the background or the MYO touches the LA exceeds a certain threshold.
    5-(1 criterion) Ratio between the minimal and maximal thickness of the MYO is below a given threshold
    6-(1 criterion) Ratio between the width of the LV and the average thickness of the MYO exceeds a certain threshold. Both width and thickness are computed as the total width of the structure at the middle-point of the embedded bounding box. The goal is to identify situations where the MYO is too thin with respect to the size of the LV.
    7-Curvature
"""
