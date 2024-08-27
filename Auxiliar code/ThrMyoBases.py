# -*- coding: utf-8 -*-
"""
Created on Thu Jul 11 15:06:13 2024

@author: RubenSilva
"""

import numpy as np
from sklearn.decomposition import PCA
import openpyxl

from CheckPlausibility import *

def ExcelRowThresolds(path,H,HT,patient):
    
    isExist = os.path.exists(path)
    if not isExist:                         
        # Create a new directory because it does not exist 
        os.makedirs(path)  
    os.chdir(path)   

    name='ThBasesMyoCropY_v2'        
    filename = str(name)+'.xlsx'
    
    print('Saving evaluation metrics to excel...')
    # Check if the file already exists
    if not os.path.isfile(filename):
        # Create a new workbook object
        book = openpyxl.Workbook()
        # Select the worksheet to add data to
        sheet = book.active
        # Add a header row to the worksheet
        sheet.append(['Patient', 'X crop','X Total'])
    else:
        # Open an existing workbook
        book = openpyxl.load_workbook(filename)
    
        # Select the worksheet to add data to
        sheet = book.active
    
    # Append a new row of data to the worksheet
    sheet.append([patient, H,HT])
    #print([patient, slic, label, pred])
    # Save the workbook to a file
    
    book.save(filename)
    
def ExcelRowThikThresolds(path,maxthik,minthik,avgthik,lv_w,patient):
    
    isExist = os.path.exists(path)
    if not isExist:                         
        # Create a new directory because it does not exist 
        os.makedirs(path)  
    os.chdir(path)   

    name='Thikness_Thrlds_ED'        
    filename = str(name)+'.xlsx'
    
    print('Saving evaluation metrics to excel...')
    # Check if the file already exists
    if not os.path.isfile(filename):
        # Create a new workbook object
        book = openpyxl.Workbook()
        # Select the worksheet to add data to
        sheet = book.active
        # Add a header row to the worksheet
        sheet.append(['Patient', 'Max Thikness','Min Thikness','Avg Thikness','LV Width'])
    else:
        # Open an existing workbook
        book = openpyxl.load_workbook(filename)
    
        # Select the worksheet to add data to
        sheet = book.active
    
    # Append a new row of data to the worksheet
    sheet.append([patient, maxthik,minthik,avgthik, lv_w])
    #print([patient, slic, label, pred])
    # Save the workbook to a file
    
    book.save(filename)    

from scipy.signal import medfilt

def RemovePointsMyo(points,x=1):
    
    """Remove the undesired points for the Myocardium"""
    
    TopRightIndex=np.argmax(points[:, x-1] + points[:, x])
    PointsToRemove_1=points[0:TopRightIndex]
    Point1=points[TopRightIndex]
    
    """Points have to be in order along the contour, find peaks in X (because the there is always a difference) """
    
    # Define the kernel size for the median filter
    kernel_size = 9

    # Apply the median filter
    der1=derivative(points[:,x])
    median_filtered_signal = medfilt(der1, kernel_size=kernel_size)
    
    nrp=len(points[:,x])//4
    
    
    # plt.plot(median_filtered_signal)
    # plt.title('Derivative in Data')
    # plt.xlabel('Index')
    # plt.ylabel('Value')
    # plt.show()
    
    
    der2=derivative(median_filtered_signal[-nrp:-1])
    minder2=np.argmin(der2)
    
    # plt.plot(der2)
    # plt.title('Second Derivative in Data')
    # plt.xlabel('Index')
    # plt.ylabel('Value')
    # plt.show()
    
    
    ind=len(points[:,x][0:-nrp])+minder2
    PointsToRemove_2=points[ind:]
    
    # Print the indices of the peaks
    #print("Indices of local maxima:", peaks)
    
    # # Print the values of the peaks
    # print("Values of local maxima:",points[:,1][peaks])
    
    # Plot the data and the peaks
    # plt.plot(points[:,x])
    # plt.plot(ind, points[:,x][ind], "x")
    # plt.title('Local Maxima in Data')
    # plt.xlabel('Index')
    # plt.ylabel('Value')
    # plt.show()
    # filtered_1=RemovePoints(points,PointsToRemove_1)
    # filtered_2=RemovePoints(filtered_1,PointsToRemove_2)
    #print(PointsToRemove_1)
    #print(PointsToRemove_2)
    
    return  Point1, points[ind],points[0],points[-1] 

def RemovePointsMyold(points,x=1):
    
    """Remove the undesired points for the Myocardium"""
    
    TopRightIndex=np.argmax(points[:, x-1] + points[:, x])
    PointsToRemove_1=points[0:TopRightIndex]
    Point1=points[TopRightIndex]
    
    #distance=AddDistancePoint(points)
    
    #plt.plot(distance,points[:,1])
    
    # Find peaks (local maxima)
    #peaks, _ = find_peaks(points[:,1])
    
    """Points have to be in order along the contour, find peaks in X (because the there is always a difference) """
    # Find peaks with a minimum height of 2 and minimum distance of 1
    peaks, properties = find_peaks(points[:,x], height=2, distance=3)
    PointsToRemove_2=points[peaks[-1]:]
    
    Point2=points[peaks[-1]]
    
    # Print the indices of the peaks
    #print("Indices of local maxima:", peaks)
    
    # # Print the values of the peaks
    # print("Values of local maxima:",points[:,1][peaks])
    
    # # Plot the data and the peaks
    # plt.plot(points[:,x])
    # plt.plot(peaks, points[:,x][peaks], "x")
    # plt.title('Local Maxima in Data')
    # plt.xlabel('Index')
    # plt.ylabel('Value')
    # plt.show()
    # filtered_1=RemovePoints(points,PointsToRemove_1)
    # filtered_2=RemovePoints(filtered_1,PointsToRemove_2)
    #print(PointsToRemove_1)
    #print(PointsToRemove_2)
    
    return  Point1,Point2, points[0],points[-1]  

from skimage.morphology import skeletonize

def compute_normal(tck, u, point_index):

    # Get the point and its derivatives
    point = np.array(splev(u[point_index], tck))
    der = np.array(splev(u[point_index], tck, der=1))
    
    normal = np.array([-der[1], der[0]])
    normal /= np.linalg.norm(normal)

    return point, normal

from scipy.interpolate import splprep, splev

def CalculateThikness(contour_coords,centerline_coords,binary_image):
    
   return 

# def RemoveSegmentsMyoTest_app3(c_pca, c_myo, seg_myo,tolerance=10, endo=1):
      
#     pca = PCA(n_components=2)
#     pca.fit(c_pca-np.mean(c_pca,axis=0))
 
#     pca_comp=pca.components_ 
      
#     if endo & (pca_comp[0,1]>0):
#         pca_comp=-pca_comp
    
#     """Myocardium rotated"""
#     c_myo_r= c_myo-np.mean(c_pca,axis=0)  
#     c_myo_r =np.matmul(c_myo_r,pca_comp.T)

#     """Calculation of the Dist transform for the mask of myocardium"""
    
#     # Compute the distance transform ###################################################################################################
 
#     dist_transform = cv2.distanceTransform(seg_myo.astype(np.uint8), cv2.DIST_L2, 5)
 
#     # Obtain the centerline by thresholding the distance transform #####################################################################
 
#     _, centerline = cv2.threshold(dist_transform, 10, 1, cv2.THRESH_BINARY)
#     centerline = np.uint8(centerline)
 
#     # Thinning the centerline to a single-pixel-wide line using skeletonize
#     centerline_skeleton = skeletonize(centerline)  # skeletonize expects binary input (0 and 1)
 
#     # Convert the centerline to coordinates
#     centerline_coords = np.column_stack(np.where(centerline_skeleton > 0))
    
#     plt.scatter(c_myo[:,1],c_myo[:,0])
#     plt.scatter(centerline_coords[:,1],centerline_coords[:,0])
#     plt.show()
 
#     plt.imshow(seg_myo)
#     plt.imshow(centerline_skeleton)
    
#     """Rotate Centerline"""
    
#     centerline_coords= centerline_coords-np.mean(c_pca,axis=0)  
#     centerline_coords =np.matmul(centerline_coords,pca_comp.T)
    
#     """Aplicar threshold"""
  
#     if endo:
#         ratio=0.097281305 # ratio= h_crop/total img
#         h_img= np.max(c_myo_r[:,0])-np.min(c_myo_r[:,0])
#         h_crop= ratio*h_img + tolerance
        
#         c_myo_r=CropY(c_myo_r, np.min(c_myo_r[:,0])+h_crop)
#         centerline_coords=CropY(centerline_coords, np.min(centerline_coords[:,0])+h_crop)
        
#     else:
#         ratio=0.070294451
#         h_img= np.max(c_myo_r[:,1])-np.min(c_myo_r[:,1])
#         h_crop= ratio*h_img + tolerance
        
#         c_myo_r=CropX(c_myo_r, np.min(c_myo_r[:,1])+h_crop)
#         centerline_coords=CropX(centerline_coords, np.min(centerline_coords[:,1])+h_crop)
    
#     #ExcelRowThresolds(path,x_crop,x_img,patient)
  
#     """Myocardium cropped in original position"""  
  
#     c_myo_r=np.matmul(c_myo_r,pca_comp)
#     c_myo_r=c_myo_r+ np.mean(c_pca,axis=0)
    
#     """Centerline with dist transform cropped in original position"""  
  
#     centerline_coords=np.matmul(centerline_coords,pca_comp)
#     centerline_coords=centerline_coords+ np.mean(c_pca,axis=0)
#     centerline_coords=centerline_coords.astype(np.int64)
    
#     plt.scatter(c_myo_r[:,1],c_myo_r[:,0])
#     # Set equal scaling
#     plt.axis('equal')
#     # plt.savefig(os.path.join(path,'CropX',patient+'.png'), format='png')
#     plt.scatter(centerline_coords[:,1],centerline_coords[:,0])
#     plt.show()
    
#     dist_centerline= dist_transform[centerline_coords[:,0], centerline_coords[:,1]]
    
#     return c_myo_r,centerline_coords,dist_centerline



# def RemoveSegmentsMyoTest_app2(c_pca, c_myo, seg_myo,tolerance=10, endo=1):
      
#     pca = PCA(n_components=2)
#     pca.fit(c_pca-np.mean(c_pca,axis=0))
 
#     pca_comp=pca.components_ 
      
#     if endo & (pca_comp[0,1]>0):
#         pca_comp=-pca_comp
    
#     """Myocardium rotated"""
#     c_myo_r= c_myo-np.mean(c_pca,axis=0)  
#     c_myo_r =np.matmul(c_myo_r,pca_comp.T)

#     """Calculation of the Dist transform for the mask of myocardium"""
    
#     dist = distance_transform_edt(seg_myo)
    
#     # Detect local maxima in the distance transform
#     local_maxima = peak_local_max(dist, min_distance=1)
    
#     """Dist transform rotated"""
#     local_maxima= local_maxima-np.mean(c_pca,axis=0)  
#     local_maxima =np.matmul(local_maxima,pca_comp.T)
    
#     plt.scatter(c_myo_r[:,1],c_myo_r[:,0])
#     # Set equal scaling
#     plt.axis('equal')
#     # plt.savefig(os.path.join(path,'CropX',patient+'.png'), format='png')
#     plt.scatter(local_maxima[:,1],local_maxima[:,0])
#     plt.show()
    
#     """Aplicar threshold"""
  
#     if endo:
#         ratio=0.097281305 # ratio= h_crop/total img
#         h_img= np.max(c_myo_r[:,0])-np.min(c_myo_r[:,0])
#         h_crop= ratio*h_img + tolerance
        
#         c_myo_r=CropY(c_myo_r, np.min(c_myo_r[:,0])+h_crop)
#         local_maxima=CropY(local_maxima, np.min(local_maxima[:,0])+h_crop)
        
#     else:
#         ratio=0.070294451
#         h_img= np.max(c_myo_r[:,1])-np.min(c_myo_r[:,1])
#         h_crop= ratio*h_img + tolerance
        
#         c_myo_r=CropX(c_myo_r, np.min(c_myo_r[:,1])+h_crop)
#         local_maxima=CropX(local_maxima, np.min(local_maxima[:,1])+h_crop)
    
#     #ExcelRowThresolds(path,x_crop,x_img,patient)
  
#     """Myocardium cropped in original position"""  
  
#     c_myo_r=np.matmul(c_myo_r,pca_comp)
#     c_myo_r=c_myo_r+ np.mean(c_pca,axis=0)
    
#     """Centerline with dist transform cropped in original position"""  
  
#     local_maxima=np.matmul(local_maxima,pca_comp)
#     local_maxima=local_maxima+ np.mean(c_pca,axis=0)
#     local_maxima=local_maxima.astype(np.int64)
    
#     plt.scatter(c_myo_r[:,1],c_myo_r[:,0])
#     # Set equal scaling
#     plt.axis('equal')
#     # plt.savefig(os.path.join(path,'CropX',patient+'.png'), format='png')
#     plt.scatter(local_maxima[:,1],local_maxima[:,0])
#     plt.show()
    
#     dist_centerline= dist[local_maxima[:,0], local_maxima[:,1]]
    
#     return c_myo_r,local_maxima,dist_centerline

def RemoveSegmentsMyoTest(c_pca, c_myo_all,c_myo,c_endo ,tolerance=25, endo=0):
      
    pca = PCA(n_components=2)
    pca.fit(c_pca-np.mean(c_pca,axis=0))
 
    pca_comp=pca.components_ 
          
    if pca_comp[1,1]>0:   
        pca_comp=-pca_comp
    
    """Myocardium rotated"""
    c_myo_all=c_myo_all-np.mean(c_pca,axis=0) 
    c_myo_all =np.matmul(c_myo_all,pca_comp.T)
    
    c_myo_r= c_myo-np.mean(c_pca,axis=0)  
    c_myo_r =np.matmul(c_myo_r,pca_comp.T)
    
    """Endocardium rotated"""
    
    c_endo_r= c_endo-np.mean(c_pca,axis=0)  
    c_endo_r =np.matmul(c_endo_r,pca_comp.T)
    
    """Apply threshold"""
      
    ratio=0.070294451
    h_img= np.max(c_myo_all[:,1])-np.min(c_myo_all[:,1])
    h_crop= ratio*h_img + tolerance
    
    c_myo_r=CropX(c_myo_r, np.min(c_myo_all[:,1])+h_crop)
    c_endo_r=CropX(c_endo_r, np.min(c_myo_all[:,1])+h_crop)
    
    LV_width= np.max(c_endo_r[:,0])-np.min(c_endo_r[:,0])
    """Myocardium cropped in original position"""  
  
    c_myo_r=np.matmul(c_myo_r,pca_comp)
    c_myo_r=c_myo_r+ np.mean(c_pca,axis=0)
    
    c_endo_r=np.matmul(c_endo_r,pca_comp)
    c_endo_r=c_endo_r+ np.mean(c_pca,axis=0)
    
    plt.scatter(c_myo_r[:,1],c_myo_r[:,0])
    plt.scatter(c_endo_r[:,1],c_endo_r[:,0])
    # Set equal scaling
    plt.axis('equal')
    plt.show()
    
    return c_myo_r, c_endo_r, LV_width

def RemoveSegmentsMyo(c_pca, c_myo,c_myo_crop, endo=1):
    
    """Remove points Myo"""
    Point1,Point2,Point3,Point4=RemovePointsMyo(c_myo_crop,x=1)
    
    
    plt.scatter(c_myo[:,1],c_myo[:,0])
    plt.plot(Point1[1],Point1[0],"x",color='r')
    plt.plot(Point2[1],Point2[0],"x",color='r')

    # Set equal scaling
    plt.axis('equal')
    plt.show()
    
    pca = PCA(n_components=2)
    pca.fit(c_pca-np.mean(c_pca,axis=0))


    contour_pca_r,c_myo_r,P1_r,P2_r,P3_r,P4_r= c_pca-np.mean(c_pca,axis=0),c_myo-np.mean(c_pca,axis=0), Point1-np.mean(c_pca,axis=0),Point2 -np.mean(c_pca,axis=0),Point3 -np.mean(c_pca,axis=0),Point4 -np.mean(c_pca,axis=0)
    
    #pca_comp=abs(pca.components_)
    #pca_comp[0,1]=-pca_comp[0,1]
    pca_comp=pca.components_ # --> Caso da Interseçao
    #pca_comp = pca_comp[:, [1, 0]] # no caso da interseção
    # pca_comp[1,0]=-pca_comp[1,0]
    #pca_comp=pca_comp.T
    
    if endo & (pca_comp[0,1]>0):
        pca_comp=-pca_comp
    
        
    contour_pca_r,c_myo_r =np.matmul(contour_pca_r,pca_comp.T), np.matmul(c_myo_r,pca_comp.T)
    P1_r,P2_r,P3_r,P4_r=np.matmul(P1_r,pca_comp.T),np.matmul(P2_r,pca_comp.T),np.matmul(P3_r,pca_comp.T),np.matmul(P4_r,pca_comp.T)
    
    plt.scatter(contour_pca_r[:,1],contour_pca_r[:,0])
    plt.scatter(c_myo_r[:,1],c_myo_r[:,0])
    plt.plot(P1_r[1],P1_r[0],"x",color='r')
    plt.plot(P2_r[1],P2_r[0],"x",color='r')
    plt.plot(P3_r[1],P3_r[0],"x",color='r')
    plt.plot(P4_r[1],P4_r[0],"x",color='r')

    # Set equal scaling
    plt.axis('equal')
    # plt.savefig(os.path.join(path,'CropX',patient+'.png'), format='png')
    plt.show()
    
    """Encontrar thresold"""
  
    if endo:
        
        x_crop=abs(-np.max([P1_r[0],P2_r[0],P3_r[0],P4_r[0]])+np.min([P1_r[0],P2_r[0],P3_r[0],P4_r[0]]))
        x_img= np.max(c_myo_r[:,0])-np.min(c_myo_r[:,0])
        c_myo_r=CropY(c_myo_r, np.min(c_myo_r[:,0])+x_crop)
        print('Ratio to crop: ',x_crop/x_img)
        
        
    else:
        
        x_crop=abs(np.max([P1_r[1],P2_r[1],P3_r[1],P4_r[1]])-np.min([P1_r[1],P2_r[1],P3_r[1],P4_r[1]]))
        x_img= np.max(c_myo_r[:,1])-np.min(c_myo_r[:,1])
        c_myo_r=CropX(c_myo_r, np.min(c_myo_r[:,1])+x_crop)
        print('Ratio to crop: ',x_crop/x_img)
    
    
    #ExcelRowThresolds(path,x_crop,x_img,patient)
  
    
    c_myo_r=np.matmul(c_myo_r,pca_comp)

    c_myo_r=c_myo_r+ np.mean(c_pca,axis=0)
    
    return c_myo_r, x_crop/x_img

from scipy.ndimage import distance_transform_edt 
from scipy.spatial import KDTree

def measure_thickness(c_myo,c_endo): 

    # Build KDTree for contour2
    tree = KDTree(c_endo)

    # Find the minimal distance from each point in contour1 to contour2
    distances, _ = tree.query(c_myo)
    plt.scatter(c_myo[:,1],c_myo[:,0])
    plt.scatter(c_endo[:,1],c_endo[:,0])
    plt.show()
    
    min_thickness = np.min(distances)
    max_thickness = np.max(distances)
    avg_thickness = np.mean(distances)
    
    
    return max_thickness, min_thickness, avg_thickness, distances 

def FindNumberContours(c,delta):
    contours=[]
    indices=[0]
    
    for i in range(len(c)-1):
        if dist2points(c[i],c[i+1])>delta:
            indices.append(i)
            
    indices.append(-1)           
            
    if len(indices)>2:
        for i in range(len(indices)-1):
            if i==0:
                contour=c[indices[i]:indices[i+1]+1]
                if len(contour)>10:
                    contours.append(contour)
            elif (i==(len(indices)-2)):
                contour=c[indices[i]+1:]
                if len(contour)>10:
                    contours.append(contour)
            else:
                contour=c[indices[i]+1:indices[i+1]+1]
                if len(contour)>10:
                    contours.append(contour)  
    else:
        
        contours.append(c)
    
    return contours  

from skimage.feature import peak_local_max
  
def FindCropTh(image,patient,path):
    
    plt.imshow(image, cmap='gray')
    plt.show()
    
    """Extract Structures"""
    
    SpecificSeg1=SeeSpecificStructure(1, image) # Endocardium
    SpecificSeg2=SeeSpecificStructure(2, image) #Myocardium
    SpecificSeg3=SeeSpecificStructure(3, image) #LA
    
    """Define contours"""
    contour_1=DefineContourSkimage(SpecificSeg1)
    contour_2=DefineContourSkimage(SpecificSeg2)
    contour_3=DefineContourSkimage(SpecificSeg3)
    
    """Interpolation to make points equisdistance"""
    
    contour_1=ContourInterpolation(contour_1)
    contour_2=ContourInterpolation(contour_2)
    contour_3=ContourInterpolation(contour_3)
    
    "Find intersections between contours"
    
    inter_13,inter_31=FindCommonPoints_new(contour_1,contour_3,3) # Endo and LA
    inter_21,_=FindCommonPoints_new(contour_2,contour_1,4) # Myo and Endo
    
    """Remove intersections from the contours"""
    new_contour_1=RemovePoints(contour_1,inter_13)
    new_contour_2=RemovePoints(contour_2,inter_21)
    new_contour_3=RemovePoints(contour_3,inter_31)
    
    """Order contour"""
    
    new_contour_1=order_contour_points(new_contour_1,2)
    new_contour_2=order_contour_points(new_contour_2,2) # fix problem with order Myo
    new_contour_3=order_contour_points(new_contour_3) # fix problem with order Atrium 

    plt.scatter(new_contour_2[:, 1], new_contour_2[:, 0], cmap='gray')
    plt.show()
    
    
    """Apply PCA to Endocardium (or intersection between and endo?)"""      
    
    """Find thresholds"""
    # contour_2_r,ratio=RemoveSegmentsMyo(inter_13, contour_2,new_contour_2, endo=0) 
   
    # plt.scatter(contour_2_r[:,1],contour_2_r[:,0])

    # # Set equal scaling
    # plt.axis('equal')
    # plt.show()       
    
    """Apply thresholds"""
    contour_myo,contour_endo,LV_width=RemoveSegmentsMyoTest(inter_13, contour_2,new_contour_2,new_contour_1)   
    
    """Find thickness thresholds"""
    
    thk_max,thk_min,thk_avg,_=measure_thickness(contour_myo,contour_endo)
    
    ExcelRowThikThresolds(path,thk_max,thk_min,thk_avg,LV_width,patient)
     
    return thk_max,LV_width    

# SofiaSegPath='C:/Users/RubenSilva/Desktop/CAMUS Dataset/SofiaSeg'
# CAMUSPath='C:/Users/RubenSilva/Desktop/CAMUS Dataset/CAMUS/patient0089'

# img_sofia = nib.load(os.path.join(SofiaSegPath,'patient0089_2CH_ES.nii.gz'))
# img_camus = nib.load(os.path.join(CAMUSPath,'patient0089_2CH_ES.nii.gz'))
# seg_camus= nib.load(os.path.join(CAMUSPath,'patient0089_2CH_ES_gt.nii.gz'))

# data_sofia= img_sofia.get_fdata()
# data_camus= img_camus.get_fdata()
# data_GT= seg_camus.get_fdata()

# th=FindCropTh(data_GT)

SofiaSegPath='C:/Users/RubenSilva/Desktop/CAMUS Dataset/SofiaSeg'
CAMUSPath='E:/RubenSilva/POCUS/Resources/database_nifti'
list_patients=os.listdir(CAMUSPath)
#list_patients=['patient0004']
#img_sofia = nib.load(os.path.join(SofiaSegPath,'patient0089_2CH_ES.nii.gz'))

"""Load images Sistole"""
path_save_info='C:/Users/RubenSilva/Desktop/THOR/findThresolds'

for patient in list_patients:
    
    path_patient= os.path.join(CAMUSPath,patient)
    img_camus = nib.load(os.path.join(path_patient,str(patient)+'_2CH_ED.nii.gz'))
    seg_camus= nib.load(os.path.join(path_patient,str(patient)+'_2CH_ED_gt.nii.gz'))

    data_camus= img_camus.get_fdata()
    data_GT= seg_camus.get_fdata()
    data_GT = np.hstack((data_GT, np.zeros((data_GT.shape[0], 1))))

    plt.imshow(data_camus, cmap='gray')
    plt.show()
    plt.imshow(data_GT, cmap='gray')
    plt.show()
    
    thk_max, lv_w=FindCropTh(data_GT,patient,path_save_info)
    print('Patient: ', patient,'thk max and LV width:', thk_max, lv_w)

# Patient 0002, 0005, 0009, 0011(estrnaho) --> Does not get vertical
# 0004, 0006, 0013 --> Invert pca components 
# 0007 --> invert

"""
array([[-0.41929523, -0.90784994],
       [ 0.90784994, -0.41929523]]) --> para baixo

array([[ 0.66137264,  0.75005749],
       [-0.75005749,  0.66137264]])--> para cima (normal)

array([[-0.51425657,  0.85763639],
       [-0.85763639, -0.51425657]]) --> para cima para certos casos

array([[-0.27394721, -0.96174473],
       [-0.96174473,  0.27394721]]) --> para baixo mas imperfeito

array([[ 0.27394721, -0.96174473],
       [ 0.96174473,  0.27394721]]) --> para baixo mas imperfeito

array([[ 0.08846766, -0.99607905],
       [ 0.99607905,  0.08846766]]) --> para baixo e perfeito

"""